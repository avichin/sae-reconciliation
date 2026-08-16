"""
SAE Reconciliation Script - Multi-Product Trial
===================================================
Variant of the standard SAE reconciliation script for trials with more
than one investigational product. Developed specifically for certain 
file formats as outlined below. 

  - The VENDOR file has one row per (SAE record, product) combination.
    The same record is duplicated across multiple rows with case-level
    fields (dates, outcome, seriousness criteria, etc.) repeating
    identically, and only the product name, Action Taken, and
    Relationship to Study Drug varying per row.
  - The EDC "Adverse Events Form" has one row per AE, with SEPARATE Action
    Taken and Relationship to Study Drug columns per product.

The output is the same two-tab structure as the standard script ("Matching
SAEs" / "Mismatching or Missing SAEs"), except Action Taken and
Relationship to Study Drug each become one column PER PRODUCT instead of
a single column.

Requirements:
    pip install pandas openpyxl
"""

import re
from datetime import datetime, timedelta
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# =========================================================================
# 1. CONFIGURATION
# =========================================================================

CONFIG = {
    "ae_file": "ae_listing.csv",              # Replace with your EDC AE file
    "vendor_file": "vendor_sae_listing.csv",  # Replace with your vendor SAE file
    "output_file": "SAE_Reconciliation_Report.xlsx",

    # Vendor file's column headers are on the 1st row of the file. Change if there are 
    # multiple header rows. 
    "vendor_header_row": 0,

    # --- EDC listing columns (case-level only; per-product
    # Action Taken / Relationship fields are configured separately below
    # under "products") ---
    # Replace with actual EDC column names
    "ae_columns": {
        "subject_id": "SUBJECT_ID_COLUMN",
        "record_position": "RECORD_POSITION_COLUMN",
        "ae_term": "AE_TERM_COLUMN",
        "ae_pt": "AE_PREFERRED_TERM_COLUMN",
        "start_date": "AE_START_DATE_COLUMN",
        "end_date": "AE_END_DATE_COLUMN",
        "serious_flag": "AE_SERIOUS_FLAG_COLUMN",
        "serious_yes_values": ["Y", "YES"],
        "ctcae_grade": "CTCAE_GRADE_COLUMN",
        "outcome": "AE_OUTCOME_COLUMN",
        "crit_death": "CRIT_DEATH_COLUMN",
        "crit_life_threatening": "CRIT_LIFE_THREATENING_COLUMN",
        "crit_hospitalization": "CRIT_HOSPITALIZATION_COLUMN",
        "crit_disability": "CRIT_DISABILITY_COLUMN",
        "crit_congenital": "CRIT_CONGENITAL_ANOMALY_COLUMN",
        "crit_other_mie": "CRIT_OTHER_MEDICALLY_IMPORTANT_COLUMN",
        "crit_other_specify": "CRIT_OTHER_SPECIFY_COLUMN",
    },

    # --- Vendor SAE listing columns (case-level; "causality" and
    # "action_taken" here are the columns read from EACH product-row) ---
    # Replace with actual vendor column names
    "vendor_columns": {
        "case_number": "VENDOR_CASE_NUMBER_COLUMN",
        "subject_id": "VENDOR_SUBJECT_ID_COLUMN",
        "product": "VENDOR_PRODUCT_COLUMN",
        "ae_term": "VENDOR_AE_TERM_COLUMN",
        "ae_pt": "VENDOR_AE_PREFERRED_TERM_COLUMN",
        "start_date": "VENDOR_START_DATE_COLUMN",
        "end_date": "VENDOR_END_DATE_COLUMN",
        "serious_flag": "VENDOR_SERIOUS_FLAG_COLUMN",
        "serious_yes_values": ["Y", "YES"],
        "ctcae_grade": "VENDOR_CTCAE_GRADE_COLUMN",
        "outcome": "VENDOR_OUTCOME_COLUMN",
        "crit_death": "VENDOR_CRIT_DEATH_COLUMN",
        "crit_life_threatening": "VENDOR_CRIT_LIFE_THREATENING_COLUMN",
        "crit_hospitalization": "VENDOR_CRIT_HOSPITALIZATION_COLUMN",
        "crit_disability": "VENDOR_CRIT_DISABILITY_COLUMN",
        "crit_congenital": "VENDOR_CRIT_CONGENITAL_ANOMALY_COLUMN",
        "crit_other_mie": "VENDOR_CRIT_OTHER_MEDICALLY_IMPORTANT_COLUMN",
        # No free-text "Other/Specify" field in this vendor file.
        "crit_other_specify": None,
        "causality": "VENDOR_CAUSALITY_COLUMN",
        "action_taken": "VENDOR_ACTION_TAKEN_COLUMN",
        # Used to filter which rows are considered at all.
        # Study ID itself is not reconciled.
        "study_id": "VENDOR_STUDY_ID_COLUMN",
    },

    # Vendor rows whose study ID doesn't match this value are dropped
    # entirely before reconciliation.
    "required_study_id": "STUDY_ID",

    # Replace with actual column names.
    "products": {
        "Product A": {"action_taken_col": "PRODUCT_A_ACTION_TAKEN_COLUMN", "causality_col": "PRODUCT_A_CAUSALITY_COLUMN"},
        "Product B": {"action_taken_col": "PRODUCT_B_ACTION_TAKEN_COLUMN", "causality_col": "PRODUCT_B_CAUSALITY_COLUMN"},
        "Product C": {"action_taken_col": "PRODUCT_C_ACTION_TAKEN_COLUMN", "causality_col": "PRODUCT_C_CAUSALITY_COLUMN"},
        "Product D": {"action_taken_col": "PRODUCT_D_ACTION_TAKEN_COLUMN", "causality_col": "PRODUCT_D_CAUSALITY_COLUMN"},
    },

    # Vendor files occasionally use different naming conventions 
    # (brand name vs. generic). Replace with the actual product names
    # and values if needed. If not needed (vendor naming matches EDC 
    # naming), either exclude or replace with the same values.
    "product_synonyms": {
        "BRAND_NAME_FOR_PRODUCT_C": "Product C",
        "BRAND_NAME_FOR_PRODUCT_D": "Product D",
    },

    "term_similarity_threshold": 0.80,
}


# =========================================================================
# 2. VALUE CROSSWALKS (Same concept as the single-product script. These
# apply per-product to Action Taken / Relationship as well as to the
# case-level CTCAE Grade / Outcome fields)
# =========================================================================

CODE_LABEL_MAPS = {
    "ctcae_grade": {
        "mild": 1, "grade 1": 1, "grade 1 mild": 1, "grade 1: mild": 1,
        "moderate": 2, "grade 2": 2, "grade 2 moderate": 2, "grade 2: moderate": 2,
        "severe": 3, "grade 3": 3, "grade 3 severe": 3, "grade 3: severe": 3,
        "life threatening": 4, "life-threatening": 4, "grade 4": 4,
        "grade 4 life-threatening": 4, "grade 4: life-threatening": 4,
        "death": 5, "grade 5": 5, "grade 5 death": 5, "grade 5: death": 5,
    },
    "outcome": {
        "fatal": 1,
        "not recovered/not resolved": 2, "not recovered": 2, "not resolved": 2,
        "recovered/resolved": 3, "recovered": 3, "resolved": 3,
        "lasting damage": 4, "recovered/resolved with sequelae": 4,
        "recovered with sequelae": 4,
        "improved": 5, "recovering/resolving": 5, "recovering": 5, "resolving": 5,
    },
    "causality": {
        "none/not reported/unknown": 1, "none": 1, "not reported": 1,
        "unknown": 1, "not applicable": 1, "n/a": 1,
        "not related": 2,
        "unlikely": 3, "unlikely related": 3,
        "possible": 4, "possibly related": 4,
        "probable": 5, "probably related": 5,
        "almost certain": 6, "definitely related": 6,
    },
    "action_taken": {
        "no change": 1, "dose not changed": 1,
        "dose decreased": 2, "dose reduced": 2,
        "dose interrupted": 3, "drug interrupted": 3,
        "withdrawn": 4, "drug withdrawn": 4, "permanently discontinued": 4,
        "not applicable": 5, "n/a": 5,
    },
}

CODED_FIELDS = ["ctcae_grade", "outcome"]  # causality/action_taken handled per-product
FLAG_FIELDS = [
    "serious_flag", "crit_death", "crit_life_threatening", "crit_hospitalization",
    "crit_disability", "crit_congenital", "crit_other_mie",
]
TEXT_FIELDS = ["ae_pt"]
DATE_FIELDS = ["start_date", "end_date"]

YES_VALUES = {"y", "yes", "1", "true", "checked", "x", "serious"}

FIELD_STATUS_LABELS = {
    "ctcae_grade": "CTCAE Grade",
    "outcome": "Outcome",
    "serious_flag": "Is the Adverse event serious (SAE)?",
    "crit_death": "Seriousness Criteria- Death",
    "crit_life_threatening": "Seriousness Criteria- Life Threatening",
    "crit_hospitalization": "Seriousness Criteria- Hospitalized",
    "crit_disability": "Seriousness Criteria- Disability",
    "crit_congenital": "Seriousness Criteria- Congenital Anomaly",
    "crit_other_mie": "Seriousness Criteria- Medically Significant",
    "ae_pt": "AE Prefered Term",
    "start_date": "Start date",
    "end_date": "End date",
}


# =========================================================================
# 3. HELPERS
# =========================================================================

def load_file(path, header=0):
    if path.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(path, dtype=str, header=header)
    else:
        df = pd.read_csv(path, dtype=str, header=header)
    df.columns = df.columns.astype(str).str.strip()
    return df


def validate_columns(df, cols, source_name):
    expected = [v for v in cols.values() if isinstance(v, str)]
    missing = [c for c in expected if c not in df.columns]
    if missing:
        raise KeyError(
            f"\n\n{source_name} file is missing expected column(s): {missing}\n\n"
            f"Columns actually found in the {source_name} file:\n  {list(df.columns)}\n\n"
            "Common causes: (1) the header row setting is wrong -- see "
            "'vendor_header_row' in CONFIG, (2) extra spaces, different "
            "capitalization, or a typo in the configured column name vs. "
            "the file, or (3) the wrong file or Excel sheet is being read."
        )


def validate_product_columns(df, products, source_name):
    expected = []
    for p, cols in products.items():
        expected.append(cols["action_taken_col"])
        expected.append(cols["causality_col"])
    missing = [c for c in expected if c not in df.columns]
    if missing:
        raise KeyError(
            f"\n\n{source_name} file is missing expected per-product column(s): "
            f"{missing}\n\nColumns actually found:\n  {list(df.columns)}"
        )


def normalize_id(x):
    if pd.isna(x):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(x)).upper()


def normalize_term(x):
    if pd.isna(x):
        return ""
    return re.sub(r"\s+", " ", str(x).strip().lower())


def parse_date(x):
    if pd.isna(x) or str(x).strip() == "":
        return pd.NaT
    s = str(x).strip()
    result = pd.to_datetime(s, format="%d%b%Y:%H:%M:%S.%f", errors="coerce")
    if pd.notna(result):
        return result
    return pd.to_datetime(s, errors="coerce")


def term_similarity(a, b):
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def _clean_text(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[^a-z0-9/ .:-]", "", s)
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_coded_value(value, field_key):
    if pd.isna(value) or str(value).strip() == "":
        return None
    raw = str(value).strip()
    cleaned = _clean_text(raw)
    m = re.match(r"^(\d+)\b[\.\-:) ]*(.*)$", cleaned)
    if m:
        return int(m.group(1))
    return CODE_LABEL_MAPS.get(field_key, {}).get(cleaned)


def normalize_flag(value):
    if pd.isna(value) or str(value).strip() == "":
        return False
    return _clean_text(value) in YES_VALUES


def canonicalize_product(raw_value, cfg):
    """Map a vendor product value to one of the canonical product
    names in CONFIG['products'], handling brand-name synonyms. Returns
    None if unrecognized."""
    if pd.isna(raw_value):
        return None
    key = str(raw_value).strip().upper()
    synonyms = {k.upper(): v for k, v in cfg["product_synonyms"].items()}
    if key in synonyms:
        return synonyms[key]
    for canonical in cfg["products"]:
        if canonical.upper() == key:
            return canonical
    return None


# =========================================================================
# 4. PREP
# =========================================================================

def prep_source(df, cols):
    out = df.copy()
    out["_subj_norm"] = out[cols["subject_id"]].apply(normalize_id)

    pt = out[cols["ae_pt"]] if cols["ae_pt"] in out else pd.Series([None] * len(out))
    verbatim = out[cols["ae_term"]]
    out["_match_term"] = pt.where(pt.notna() & (pt.astype(str).str.strip() != ""), verbatim)
    out["_match_term_norm"] = out["_match_term"].apply(normalize_term)

    out["_start_date_norm"] = out[cols["start_date"]].apply(parse_date)
    return out


def filter_serious(ae_df, cols):
    is_serious = ae_df[cols["serious_flag"]].astype(str).str.upper().isin(
        [v.upper() for v in cols["serious_yes_values"]]
    )
    return ae_df[is_serious].reset_index(drop=True)


def build_vendor_cases(vendor_df, cfg):
    """Collapse the vendor file's one-row-per-product rows into one row
    per SAE case, with a nested per-product dict of {action_taken,
    causality}. Case-level fields are taken from the first row in each
    group (they're expected to repeat identically across the group)."""
    vc = cfg["vendor_columns"]
    cases = []
    for case_num, group in vendor_df.groupby(vc["case_number"]):
        rep = group.iloc[0].to_dict()
        product_values = {}
        for _, r in group.iterrows():
            canon = canonicalize_product(r.get(vc["product"]), cfg)
            if canon is None:
                print(f"WARNING: unrecognized product '{r.get(vc['product'])}' "
                      f"for vendor case {case_num} -- skipped.")
                continue
            product_values[canon] = {
                "action_taken": r.get(vc["action_taken"]),
                "causality": r.get(vc["causality"]),
            }
        rep["_product_values"] = product_values
        cases.append(rep)
    return pd.DataFrame(cases)


# =========================================================================
# 5. FIELD-LEVEL COMPARISON
# =========================================================================

def compare_matched_pair(ae_row, v_case, ac, vc, cfg):
    """Returns a list of human-readable mismatch labels."""
    mismatched = []

    for field in CODED_FIELDS:
        ae_code = normalize_coded_value(ae_row.get(ac[field]), field)
        v_code = normalize_coded_value(v_case.get(vc[field]), field)
        if ae_code is not None and v_code is not None:
            if ae_code != v_code:
                mismatched.append(FIELD_STATUS_LABELS.get(field, field))
        elif _clean_text(ae_row.get(ac[field])) != _clean_text(v_case.get(vc[field])):
            mismatched.append(FIELD_STATUS_LABELS.get(field, field))

    for field in FLAG_FIELDS:
        if normalize_flag(ae_row.get(ac[field])) != normalize_flag(v_case.get(vc[field])):
            mismatched.append(FIELD_STATUS_LABELS.get(field, field))

    for field in TEXT_FIELDS:
        if _clean_text(ae_row.get(ac[field])) != _clean_text(v_case.get(vc[field])):
            mismatched.append(FIELD_STATUS_LABELS.get(field, field))

    for field in DATE_FIELDS:
        ae_d = parse_date(ae_row.get(ac[field]))
        v_d = parse_date(v_case.get(vc[field]))
        if pd.notna(ae_d) and pd.notna(v_d):
            if ae_d.date() != v_d.date():
                mismatched.append(FIELD_STATUS_LABELS.get(field, field))
        elif pd.notna(ae_d) != pd.notna(v_d):
            mismatched.append(FIELD_STATUS_LABELS.get(field, field))

    # Per-product Action Taken / Relationship to Study Drug.
    # If the vendor doesn't list a given product for an SAE,
    # that product is not reconciled. Skip it regardless of
    # what EDC has for it.
    product_values = v_case.get("_product_values", {}) or {}
    for product, pcols in cfg["products"].items():
        v_pv = product_values.get(product)
        if v_pv is None:
            continue

        edc_action = ae_row.get(pcols["action_taken_col"])
        edc_causality = ae_row.get(pcols["causality_col"])
        ae_action_code = normalize_coded_value(edc_action, "action_taken")
        ae_causality_code = normalize_coded_value(edc_causality, "causality")

        v_action = v_pv.get("action_taken")
        v_causality = v_pv.get("causality")
        v_action_code = normalize_coded_value(v_action, "action_taken")
        v_causality_code = normalize_coded_value(v_causality, "causality")

        if ae_action_code is not None and v_action_code is not None:
            if ae_action_code != v_action_code:
                mismatched.append(f"Action Taken - {product}")
        elif _clean_text(edc_action) != _clean_text(v_action):
            mismatched.append(f"Action Taken - {product}")

        if ae_causality_code is not None and v_causality_code is not None:
            if ae_causality_code != v_causality_code:
                mismatched.append(f"Relationship to Study Drug - {product}")
        elif _clean_text(edc_causality) != _clean_text(v_causality):
            mismatched.append(f"Relationship to Study Drug - {product}")

    return mismatched


def build_status_string(mismatched_labels):
    return ", ".join(f"{label} mismatch" for label in mismatched_labels)


# =========================================================================
# 6. ROW BUILDERS
# =========================================================================

CASE_LEVEL_FIELD_MAP = [
    ("AE Term", "ae_term"),
    ("AE Prefered Term", "ae_pt"),
    ("Is the Adverse event serious (SAE)?", "serious_flag"),
    ("Start Date", "start_date"),
    ("End Date", "end_date"),
    ("CTCAE Grade", "ctcae_grade"),
    ("Outcome", "outcome"),
    ("Death", "crit_death"),
    ("Life Threatening", "crit_life_threatening"),
    ("Initial or Prolonged Hospitalization", "crit_hospitalization"),
    ("Disability/Incapacity", "crit_disability"),
    ("Congenital Anomoly/Birth Defect", "crit_congenital"),
    ("Other Medically Important Event", "crit_other_mie"),
    ("Other Medically Important Event (Specify)", "crit_other_specify"),
]


def product_column_labels(cfg):
    labels = []
    for product in cfg["products"]:
        labels.append(f"Action Taken - {product}")
        labels.append(f"Relationship to Study Drug - {product}")
    return labels


def build_tab_columns(id_cols, cfg):
    return id_cols + [lbl for lbl, _ in CASE_LEVEL_FIELD_MAP] + product_column_labels(cfg)


def build_case_level_dict(row, cols):
    d = {}
    for label, key in CASE_LEVEL_FIELD_MAP:
        col_name = cols.get(key)
        d[label] = row.get(col_name) if col_name else None
    return d


def build_edc_product_dict(ae_row, cfg):
    d = {}
    for product, pcols in cfg["products"].items():
        d[f"Action Taken - {product}"] = ae_row.get(pcols["action_taken_col"])
        d[f"Relationship to Study Drug - {product}"] = ae_row.get(pcols["causality_col"])
    return d


def build_vendor_product_dict(v_case, cfg):
    product_values = v_case.get("_product_values", {}) or {}
    d = {}
    for product in cfg["products"]:
        pv = product_values.get(product, {})
        d[f"Action Taken - {product}"] = pv.get("action_taken")
        d[f"Relationship to Study Drug - {product}"] = pv.get("causality")
    return d


def build_edc_row(ae_row, ac, status, cfg, vendor_case_number=None):
    d = {
        "Subject Number": ae_row.get(ac["subject_id"]),
        "Vendor Case #": vendor_case_number,
        "EDC Row #": ae_row.get(ac.get("record_position")),
        "Source": "EDC",
        "Status": status,
    }
    d.update(build_case_level_dict(ae_row, ac))
    d.update(build_edc_product_dict(ae_row, cfg))
    return d


def build_vendor_row(v_case, vc, status, cfg, edc_row_num=None):
    d = {
        "Subject Number": v_case.get(vc["subject_id"]),
        "Vendor Case #": v_case.get(vc.get("case_number")),
        "EDC Row #": edc_row_num,
        "Source": "Vendor",
        "Status": status,
    }
    d.update(build_case_level_dict(v_case, vc))
    d.update(build_vendor_product_dict(v_case, cfg))
    return d


def build_merged_row(ae_row, v_case, ac, vc, cfg):
    """EDC values are source of truth for a fully-matching case; vendor
    case number carried through for traceability."""
    d = {
        "Subject Number": ae_row.get(ac["subject_id"]),
        "Case #": v_case.get(vc.get("case_number")),
        "EDC Row #": ae_row.get(ac.get("record_position")),
    }
    d.update(build_case_level_dict(ae_row, ac))
    d.update(build_edc_product_dict(ae_row, cfg))
    return d


# =========================================================================
# 7. RECONCILIATION
# =========================================================================

def reconcile(sae_df, vendor_cases_df, cfg):
    ac, vc = cfg["ae_columns"], cfg["vendor_columns"]
    threshold = cfg["term_similarity_threshold"]
    LARGE_GAP = timedelta(days=999999)

    matching_rows, mismatch_rows = [], []
    used_vendor_idx = set()

    for key, sae_group in sae_df.groupby("_subj_norm"):
        vendor_group = vendor_cases_df[vendor_cases_df["_subj_norm"] == key]

        for _, ae_row in sae_group.iterrows():
            candidates = []
            for v_idx, v_case in vendor_group.iterrows():
                if v_idx in used_vendor_idx:
                    continue
                score = term_similarity(ae_row["_match_term_norm"], v_case["_match_term_norm"])
                if score < threshold:
                    continue
                date_diff = LARGE_GAP
                if pd.notna(ae_row["_start_date_norm"]) and pd.notna(v_case["_start_date_norm"]):
                    date_diff = abs(ae_row["_start_date_norm"] - v_case["_start_date_norm"])
                candidates.append((score, date_diff, v_idx, v_case))

            if candidates:
                candidates.sort(key=lambda c: (-round(c[0], 4), c[1]))
                _, _, v_idx, v_case = candidates[0]
                used_vendor_idx.add(v_idx)

                mismatched = compare_matched_pair(ae_row, v_case, ac, vc, cfg)
                if mismatched:
                    status = build_status_string(mismatched)
                    edc_row_num = ae_row.get(ac.get("record_position"))
                    vendor_case_num = v_case.get(vc.get("case_number"))
                    mismatch_rows.append(build_edc_row(ae_row, ac, status, cfg, vendor_case_number=vendor_case_num))
                    mismatch_rows.append(build_vendor_row(v_case, vc, status, cfg, edc_row_num=edc_row_num))
                else:
                    matching_rows.append(build_merged_row(ae_row, v_case, ac, vc, cfg))
            else:
                mismatch_rows.append(build_edc_row(ae_row, ac, "Missing in vendor", cfg))

    for v_idx, v_case in vendor_cases_df.iterrows():
        if v_idx not in used_vendor_idx:
            mismatch_rows.append(build_vendor_row(v_case, vc, "Missing in EDC", cfg))

    return matching_rows, mismatch_rows


# =========================================================================
# 8. MAIN
# =========================================================================

def wrap_comments_column(path, sheet_name, column_label, width=45):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]
    if column_label not in header:
        wb.close()
        return
    col_idx = header.index(column_label) + 1
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    ws.column_dimensions[col_letter].width = width
    for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)


def main():
    cfg = CONFIG
    ac, vc = cfg["ae_columns"], cfg["vendor_columns"]

    ae_raw = load_file(cfg["ae_file"])
    vendor_raw = load_file(cfg["vendor_file"], header=cfg["vendor_header_row"])

    validate_columns(ae_raw, ac, "EDC")
    validate_columns(vendor_raw, vc, "vendor")
    validate_product_columns(ae_raw, cfg["products"], "EDC")

    if vc.get("study_id") and cfg.get("required_study_id"):
        required = str(cfg["required_study_id"]).strip()
        before = len(vendor_raw)
        vendor_raw = vendor_raw[
            vendor_raw[vc["study_id"]].astype(str).str.strip() == required
        ].reset_index(drop=True)
        excluded = before - len(vendor_raw)
        if excluded:
            print(f"Excluded {excluded} vendor row(s) not belonging to study "
                  f"'{required}' (out of {before} total).")

    ae_prepped = prep_source(ae_raw, ac)
    sae_df = filter_serious(ae_prepped, ac)

    vendor_cases = build_vendor_cases(vendor_raw, cfg)
    vendor_cases_df = prep_source(vendor_cases, vc)

    matching_rows, mismatch_rows = reconcile(sae_df, vendor_cases_df, cfg)

    match_cols = build_tab_columns(["Subject Number", "Case #", "EDC Row #"], cfg)
    mismatch_cols = build_tab_columns(["Subject Number", "Vendor Case #", "EDC Row #", "Source", "Status"], cfg)

    matching_df = pd.DataFrame(matching_rows, columns=match_cols)
    matching_df.insert(0, "Obs", range(1, len(matching_df) + 1))

    mismatch_df = pd.DataFrame(mismatch_rows, columns=mismatch_cols)
    mismatch_df.insert(0, "Obs", range(1, len(mismatch_df) + 1))

    comments_col = datetime.now().strftime("%b%Y").upper() + " DM Comments"
    mismatch_df[comments_col] = None

    with pd.ExcelWriter(cfg["output_file"], engine="openpyxl") as writer:
        matching_df.to_excel(writer, sheet_name="Matching SAEs", index=False)
        mismatch_df.to_excel(writer, sheet_name="Mismatching or Missing SAEs", index=False)

    wrap_comments_column(cfg["output_file"], "Mismatching or Missing SAEs", comments_col)

    print(f"Total serious AEs in EDC listing: {len(sae_df)}")
    print(f"Total vendor cases (after collapsing per-product rows): {len(vendor_cases_df)}")
    print(f"Matching SAEs: {len(matching_df)}")
    print(f"Rows in Mismatching or Missing SAEs tab: {len(mismatch_df)}")
    print(f"\nReport saved to {cfg['output_file']}")


if __name__ == "__main__":
    main()
