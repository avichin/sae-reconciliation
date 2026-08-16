"""
SAE Reconciliation Script
==========================
Compares Serious Adverse Events (SAEs) in the EDC AE (Adverse Events) form 
listing against the vendor's SAE listing.

Output is a single Excel workbook with two tabs:

  "Matching SAEs"
      One row per SAE where the EDC and vendor records were matched and
      every compared field agrees.

  "Mismatching or Missing SAEs"
      - For a matched SAE where one or more fields disagree: TWO rows are
        written (the vendor's version, then the EDC's version) so they can
        be visually compared line by line. Vendor rows have a blank
        "EDC Row #"; EDC rows have a blank "Vendor Case #".
      - For an SAE in the EDC listing with no vendor match: one EDC-only row.
      - For a vendor record with no matching EDC SAE: one vendor-only row.

Requirements:
    pip install pandas openpyxl
"""

import re
from datetime import datetime, timedelta
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# =========================================================================
# 1. CONFIGURATION
# =========================================================================

CONFIG = {
    "ae_file": "ae_listing.csv",              # Replace with your EDC AE file
    "vendor_file": "vendor_sae_listing.csv",  # Replace with your vendor SAE file
    "output_file": "SAE_Reconciliation_Report.xlsx",

    # Vendor file's column headers are on the 1st row of the file. Change if there are 
    # multiple header rows. 
    "vendor_header_row": 0,

    # --- EDC listing columns ---
    # Replace with actual EDC column names
    "ae_columns": {
        "subject_id": "SUBJECT_ID_COLUMN",
        # Created for EDC systems that do not allow for case numbers. If the case number
        # value exists and can be reconciled against vendor-assigned case number,
        # include it here.
        "record_position": "RECORD_POSITION_COLUMN",   
        "ae_term": "AE_TERM_COLUMN",
        "ae_pt": "AE_PREFERRED_TERM_COLUMN",
        "start_date": "AE_START_DATE_COLUMN",
        "end_date": "AE_END_DATE_COLUMN",
        "serious_flag": "AE_SERIOUS_FLAG_COLUMN",
        "serious_yes_values": ["Y", "YES"],
        "ctcae_grade": "CTCAE_GRADE_COLUMN",
        "outcome": "AE_OUTCOME_COLUMN",
        "crit_death": "CRIT_DEATH_COLUMN",
        "crit_life_threatening": "CRIT_LIFE_THREATENING_COLUMN",
        "crit_hospitalization": "CRIT_HOSPITALIZATION_COLUMN",
        "crit_disability": "CRIT_DISABILITY_COLUMN",
        "crit_congenital": "CRIT_CONGENITAL_ANOMALY_COLUMN",
        "crit_other_mie": "CRIT_OTHER_MEDICALLY_IMPORTANT_COLUMN",
        "crit_other_specify": "CRIT_OTHER_SPECIFY_COLUMN",
        "causality": "CAUSALITY_COLUMN",
        "action_taken": "ACTION_TAKEN_COLUMN",
    },

    # --- Vendor listing columns ---
    # Replace with actual vendor column names
    "vendor_columns": {
        # Vendor-assigned case number. 
        "case_number": "VENDOR_CASE_NUMBER_COLUMN",
        "subject_id": "VENDOR_SUBJECT_ID_COLUMN",
        "ae_term": "VENDOR_AE_TERM_COLUMN",
        "ae_pt": "VENDOR_AE_PREFERRED_TERM_COLUMN",
        "start_date": "VENDOR_START_DATE_COLUMN",
        "end_date": "VENDOR_END_DATE_COLUMN",
        "serious_flag": "VENDOR_SERIOUS_FLAG_COLUMN",
        "serious_yes_values": ["Y", "YES"],
        "ctcae_grade": "VENDOR_CTCAE_GRADE_COLUMN",
        "outcome": "VENDOR_OUTCOME_COLUMN",
        "crit_death": "VENDOR_CRIT_DEATH_COLUMN",
        "crit_life_threatening": "VENDOR_CRIT_LIFE_THREATENING_COLUMN",
        "crit_hospitalization": "VENDOR_CRIT_HOSPITALIZATION_COLUMN",
        "crit_disability": "VENDOR_CRIT_DISABILITY_COLUMN",
        "crit_congenital": "VENDOR_CRIT_CONGENITAL_ANOMALY_COLUMN",
        "crit_other_mie": "VENDOR_CRIT_OTHER_MEDICALLY_IMPORTANT_COLUMN",
        "crit_other_specify": "VENDOR_CRIT_OTHER_SPECIFY_COLUMN",
        "causality": "VENDOR_CAUSALITY_COLUMN",
        "action_taken": "VENDOR_ACTION_TAKEN_COLUMN",
    },

    # --- Matching tolerance ---
    # Two records are considered the SAME SAE based on Subject + AE
    # Preferred Term similarity alone. Start Date is intentionally NOT a
    # gating criterion, it is a common discrepancy for records to have 
    # mismatching start dates.
    # The Start Date is only used to disambiguate when a subject has more 
    # than one similarly-worded AE record on the vendor side.
    "term_similarity_threshold": 0.80,   # 0-1 cutoff for fuzzy AE term match
}


# =========================================================================
# 2. VALUE CROSSWALKS
# Vendor and EDC use different labels for the same fields. Update this portion
# according to the specific vendor and EDC specification.
# =========================================================================

CODE_LABEL_MAPS = {
    "ctcae_grade": {
        "mild": 1, "grade 1": 1, "grade 1 mild": 1, "grade 1: mild": 1,
        "moderate": 2, "grade 2": 2, "grade 2 moderate": 2, "grade 2: moderate": 2,
        "severe": 3, "grade 3": 3, "grade 3 severe": 3, "grade 3: severe": 3,
        "life threatening": 4, "life-threatening": 4, "grade 4": 4,
        "grade 4 life-threatening": 4, "grade 4: life-threatening": 4,
        "death": 5, "grade 5": 5, "grade 5 death": 5, "grade 5: death": 5,
    },
    "outcome": {
        "fatal": 1,
        "not recovered/not resolved": 2, "not recovered": 2, "not resolved": 2,
        "recovered/resolved": 3, "recovered": 3, "resolved": 3,
        "lasting damage": 4, "recovered/resolved with sequelae": 4,
        "recovered with sequelae": 4,
        "improved": 5, "recovering/resolving": 5, "recovering": 5, "resolving": 5,
    },
    "causality": {
        "none/not reported/unknown": 1, "none": 1, "not reported": 1,
        "unknown": 1, "not applicable": 1, "n/a": 1,
        "not related": 2,
        "unlikely": 3, "unlikely related": 3,
        "possible": 4, "possibly related": 4,
        "probable": 5, "probably related": 5,
        "almost certain": 6, "definitely related": 6,
    },
    "action_taken": {
        "no change": 1, "dose not changed": 1,
        "dose decreased": 2, "dose reduced": 2,
        "dose interrupted": 3, "drug interrupted": 3,
        "withdrawn": 4, "drug withdrawn": 4,
        "not applicable": 5, "n/a": 5,
    },
}

CODED_FIELDS = ["ctcae_grade", "outcome", "causality", "action_taken"]
FLAG_FIELDS = [
    "serious_flag", "crit_death", "crit_life_threatening", "crit_hospitalization",
    "crit_disability", "crit_congenital", "crit_other_mie",
]
TEXT_FIELDS = ["ae_pt"]
DATE_FIELDS = ["start_date", "end_date"]

YES_VALUES = {"y", "yes", "1", "true", "checked", "x", "serious"}

# The mismatching SAEs will be labeled by a human-readable Status that lists the 
# discrepancies found.  
FIELD_STATUS_LABELS = {
    "ctcae_grade": "CTCAE Grade",
    "outcome": "Outcome",
    "causality": "Relationship to Study Drug",
    "action_taken": "Action Taken with Study Drug",
    "serious_flag": "Is the Adverse event serious (SAE)?",
    "crit_death": "Seriousness Criteria- Death",
    "crit_life_threatening": "Seriousness Criteria- Life Threatening",
    "crit_hospitalization": "Seriousness Criteria- Hospitalized",
    "crit_disability": "Seriousness Criteria- Disability",
    "crit_congenital": "Seriousness Criteria- Congenital Anomaly",
    "crit_other_mie": "Seriousness Criteria- Medically Significant",
    "ae_pt": "AE Prefered Term",
    "start_date": "Start date",
    "end_date": "End date",
}


# =========================================================================
# 3. HELPERS
# =========================================================================

def load_file(path, header=0):
    if path.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(path, dtype=str, header=header)
    else:
        df = pd.read_csv(path, dtype=str, header=header)
    df.columns = df.columns.astype(str).str.strip()
    return df


def validate_columns(df, cols, source_name):
    """Fail fast with a clear, actionable message instead of a bare
    KeyError deep inside pandas if a configured column name doesn't
    actually exist in the loaded file."""
    expected = [v for v in cols.values() if isinstance(v, str)]
    missing = [c for c in expected if c not in df.columns]
    if missing:
        raise KeyError(
            f"\n\n{source_name} file is missing expected column(s): {missing}\n\n"
            f"Columns actually found in the {source_name} file:\n  "
            f"{list(df.columns)}\n\n"
            "Common causes: (1) the header row setting is wrong -- see "
            "'vendor_header_row' in CONFIG, (2) extra spaces, different "
            "capitalization, or a typo in the configured column name vs. "
            "the file, or (3) the wrong file or Excel sheet is being read."
        )


def normalize_id(x):
    if pd.isna(x):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", str(x)).upper()


def normalize_term(x):
    if pd.isna(x):
        return ""
    return re.sub(r"\s+", " ", str(x).strip().lower())


def parse_date(x):
    if pd.isna(x) or str(x).strip() == "":
        return pd.NaT
    s = str(x).strip()
    # EDC dates come as SAS-style strings, e.g. '19JAN2025:00:00:00.000'
    result = pd.to_datetime(s, format="%d%b%Y:%H:%M:%S.%f", errors="coerce")
    if pd.notna(result):
        return result
    return pd.to_datetime(s, errors="coerce")


def term_similarity(a, b):
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def _clean_text(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[^a-z0-9/ .:-]", "", s)
    s = re.sub(r"\s*/\s*", "/", s)  
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_coded_value(value, field_key):
    if pd.isna(value) or str(value).strip() == "":
        return None
    raw = str(value).strip()
    cleaned = _clean_text(raw)

    m = re.match(r"^(\d+)\b[\.\-:) ]*(.*)$", cleaned)
    if m:
        return int(m.group(1))

    return CODE_LABEL_MAPS.get(field_key, {}).get(cleaned)


def normalize_flag(value):
    """Blank/unchecked is treated as No, per typical EDC checkbox behavior."""
    if pd.isna(value) or str(value).strip() == "":
        return False
    return _clean_text(value) in YES_VALUES


# =========================================================================
# 4. PREP
# =========================================================================

def prep_source(df, cols):
    out = df.copy()
    out["_subj_norm"] = out[cols["subject_id"]].apply(normalize_id)

    pt = out[cols["ae_pt"]] if cols["ae_pt"] in out else pd.Series([None] * len(out))
    verbatim = out[cols["ae_term"]]
    out["_match_term"] = pt.where(pt.notna() & (pt.astype(str).str.strip() != ""), verbatim)
    out["_match_term_norm"] = out["_match_term"].apply(normalize_term)

    out["_start_date_norm"] = out[cols["start_date"]].apply(parse_date)
    return out


def filter_serious(ae_df, cols):
    is_serious = ae_df[cols["serious_flag"]].astype(str).str.upper().isin(
        [v.upper() for v in cols["serious_yes_values"]]
    )
    return ae_df[is_serious].reset_index(drop=True)


# =========================================================================
# 5. FIELD-LEVEL COMPARISON (decides Matching vs Mismatching)
# =========================================================================

def compare_matched_pair(ae_row, v_row, ac, vc):
    """Returns a list of field keys where the EDC and vendor values disagree
    (empty list = fully matching). Fields not mapped on both sides (e.g. a
    field only tracked in one system) are skipped rather than compared."""
    mismatched_fields = []

    def both_mapped(field):
        return ac.get(field) and vc.get(field)

    for field in CODED_FIELDS:
        if not both_mapped(field):
            continue
        ae_code = normalize_coded_value(ae_row.get(ac[field]), field)
        v_code = normalize_coded_value(v_row.get(vc[field]), field)
        if ae_code is not None and v_code is not None:
            if ae_code != v_code:
                mismatched_fields.append(field)
        elif _clean_text(ae_row.get(ac[field])) != _clean_text(v_row.get(vc[field])):
            mismatched_fields.append(field)

    for field in FLAG_FIELDS:
        if not both_mapped(field):
            continue
        if normalize_flag(ae_row.get(ac[field])) != normalize_flag(v_row.get(vc[field])):
            mismatched_fields.append(field)

    for field in TEXT_FIELDS:
        if not both_mapped(field):
            continue
        if _clean_text(ae_row.get(ac[field])) != _clean_text(v_row.get(vc[field])):
            mismatched_fields.append(field)

    for field in DATE_FIELDS:
        if not both_mapped(field):
            continue
        ae_d = parse_date(ae_row.get(ac[field]))
        v_d = parse_date(v_row.get(vc[field]))
        if pd.notna(ae_d) and pd.notna(v_d):
            if ae_d.date() != v_d.date():
                mismatched_fields.append(field)
        elif pd.notna(ae_d) != pd.notna(v_d):
            mismatched_fields.append(field)

    return mismatched_fields


def build_status_string(mismatched_fields):
    labels = [FIELD_STATUS_LABELS.get(f, f) for f in mismatched_fields]
    return ", ".join(f"{label} mismatch" for label in labels)


# =========================================================================
# 6. ROW BUILDERS FOR THE TWO OUTPUT TABS
# =========================================================================

CONTENT_FIELD_MAP = [
    ("AE Term", "ae_term"),
    ("AE Prefered Term", "ae_pt"),
    ("Is the Adverse event serious (SAE)?", "serious_flag"),
    ("Start Date", "start_date"),
    ("End Date", "end_date"),
    ("CTCAE Grade", "ctcae_grade"),
    ("Outcome", "outcome"),
    ("Death", "crit_death"),
    ("Life Threatening", "crit_life_threatening"),
    ("Initial or Prolonged Hospitalization", "crit_hospitalization"),
    ("Disability/Incapacity", "crit_disability"),
    ("Congenital Anomoly/Birth Defect", "crit_congenital"),
    ("Other Medically Important Event", "crit_other_mie"),
    ("Other Medically Important Event (Specify)", "crit_other_specify"),
    ("Relationship to Study Drug", "causality"),
    ("Action Taken with Study Drug", "action_taken"),
]

MATCH_TAB_COLUMNS = (
    ["Obs", "Subject Number", "Case #", "EDC Row #"]
    + [label for label, _ in CONTENT_FIELD_MAP]
)
MISMATCH_TAB_COLUMNS = (
    ["Obs", "Subject Number", "Vendor Case #", "EDC Row #", "Source", "Status"]
    + [label for label, _ in CONTENT_FIELD_MAP]
)


def build_content_dict(row, cols):
    d = {}
    for label, key in CONTENT_FIELD_MAP:
        col_name = cols.get(key)
        d[label] = row.get(col_name) if col_name else None
    return d


def build_edc_row(ae_row, ac, status, vendor_case_number=None):
    d = {
        "Subject Number": ae_row.get(ac["subject_id"]),
        "Vendor Case #": vendor_case_number,
        "EDC Row #": ae_row.get(ac.get("record_position")),
        "Source": "EDC",
        "Status": status,
    }
    d.update(build_content_dict(ae_row, ac))
    return d


def build_vendor_row(v_row, vc, status, edc_row_num=None):
    d = {
        "Subject Number": v_row.get(vc["subject_id"]),
        "Vendor Case #": v_row.get(vc.get("case_number")),
        "EDC Row #": edc_row_num,
        "Source": "Vendor",
        "Status": status,
    }
    d.update(build_content_dict(v_row, vc))
    return d


def build_merged_row(ae_row, v_row, ac, vc):
    """A fully matching pair: EDC values are treated as source of truth
    for the display fields, vendor's case number is carried through."""
    d = {
        "Subject Number": ae_row.get(ac["subject_id"]),
        "Case #": v_row.get(vc.get("case_number")),
        "EDC Row #": ae_row.get(ac.get("record_position")),
    }
    d.update(build_content_dict(ae_row, ac))
    return d


# =========================================================================
# 7. RECONCILIATION
# =========================================================================

def reconcile(sae_df, vendor_df, cfg):
    ac, vc = cfg["ae_columns"], cfg["vendor_columns"]
    threshold = cfg["term_similarity_threshold"]
    LARGE_GAP = timedelta(days=999999)  # sentinel for missing dates in tie-break sort

    matching_rows = []
    mismatch_rows = []
    used_vendor_idx = set()

    for key, sae_group in sae_df.groupby("_subj_norm"):
        vendor_group = vendor_df[vendor_df["_subj_norm"] == key]

        for _, ae_row in sae_group.iterrows():
            candidates = []
            for v_idx, v_row in vendor_group.iterrows():
                if v_idx in used_vendor_idx:
                    continue

                score = term_similarity(ae_row["_match_term_norm"], v_row["_match_term_norm"])
                if score < threshold:
                    continue

                date_diff = LARGE_GAP
                if pd.notna(ae_row["_start_date_norm"]) and pd.notna(v_row["_start_date_norm"]):
                    date_diff = abs(ae_row["_start_date_norm"] - v_row["_start_date_norm"])

                candidates.append((score, date_diff, v_idx, v_row))

            if candidates:
                # Best term match wins; closer start date breaks ties among
                # equally-worded records. Date is never a reason to fail
                # a match outright.
                candidates.sort(key=lambda c: (-round(c[0], 4), c[1]))
                _, _, v_idx, v_row = candidates[0]
                used_vendor_idx.add(v_idx)

                mismatched_fields = compare_matched_pair(ae_row, v_row, ac, vc)
                if mismatched_fields:
                    status = build_status_string(mismatched_fields)
                    # Both records genuinely exist and are linked.
                    # Both Case # and EDC Row # are listed on BOTH rows.
                    edc_row_num = ae_row.get(ac.get("record_position"))
                    vendor_case_num = v_row.get(vc.get("case_number"))
                    mismatch_rows.append(build_edc_row(ae_row, ac, status, vendor_case_number=vendor_case_num))
                    mismatch_rows.append(build_vendor_row(v_row, vc, status, edc_row_num=edc_row_num))
                else:
                    matching_rows.append(build_merged_row(ae_row, v_row, ac, vc))
            else:
                mismatch_rows.append(build_edc_row(ae_row, ac, "Missing in vendor"))

    for v_idx, v_row in vendor_df.iterrows():
        if v_idx not in used_vendor_idx:
            mismatch_rows.append(build_vendor_row(v_row, vc, "Missing in EDC"))

    return matching_rows, mismatch_rows


# =========================================================================
# 8. MAIN
# =========================================================================

def wrap_comments_column(path, sheet_name, *column_labels, width=45):
    """Enable wrap-text on the given column(s) (by header label) so long
    comments display on multiple lines instead of overflowing, and give
    the column a readable width to wrap at."""
    wb = load_workbook(path)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]

    for label in column_labels:
        if label not in header:
            continue
        col_idx = header.index(label) + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width
        for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def main():
    cfg = CONFIG
    ac, vc = cfg["ae_columns"], cfg["vendor_columns"]

    ae_raw = load_file(cfg["ae_file"])
    vendor_raw = load_file(cfg["vendor_file"], header=cfg["vendor_header_row"])

    validate_columns(ae_raw, ac, "EDC")
    validate_columns(vendor_raw, vc, "vendor")

    ae_prepped = prep_source(ae_raw, ac)
    sae_df = filter_serious(ae_prepped, ac)
    vendor_df = prep_source(vendor_raw, vc)

    matching_rows, mismatch_rows = reconcile(sae_df, vendor_df, cfg)

    matching_df = pd.DataFrame(matching_rows, columns=MATCH_TAB_COLUMNS[1:])
    matching_df.insert(0, "Obs", range(1, len(matching_df) + 1))
    matching_df = matching_df.reindex(columns=MATCH_TAB_COLUMNS)

    mismatch_df = pd.DataFrame(mismatch_rows, columns=MISMATCH_TAB_COLUMNS[1:])
    mismatch_df.insert(0, "Obs", range(1, len(mismatch_df) + 1))
    mismatch_df = mismatch_df.reindex(columns=MISMATCH_TAB_COLUMNS)

    # Blank column for reviewers to add comments, labeled with the month
    # this reconciliation was run (e.g. "AUG2026 DM Comments").
    comments_col = datetime.now().strftime("%b%Y").upper() + " DM Comments"
    mismatch_df[comments_col] = None

    with pd.ExcelWriter(cfg["output_file"], engine="openpyxl") as writer:
        matching_df.to_excel(writer, sheet_name="Matching SAEs", index=False)
        mismatch_df.to_excel(writer, sheet_name="Mismatching or Missing SAEs", index=False)

    wrap_comments_column(cfg["output_file"], "Mismatching or Missing SAEs", comments_col)

    print(f"Total serious AEs in EDC listing: {len(sae_df)}")
    print(f"Total records in vendor listing: {len(vendor_df)}")
    print(f"Matching SAEs: {len(matching_df)}")
    print(f"Rows in Mismatching or Missing SAEs tab: {len(mismatch_df)}")
    print("  (each field-level discrepancy = 2 rows; missing/extra = 1 row)")
    print(f"\nReport saved to {cfg['output_file']}")


if __name__ == "__main__":
    main()
