"""
SAE Reconciliation Output Comparison Script
=============================================
Compares a newly generated SAE Reconciliation output against the previous
run's output. Additionally does the following:

  1. Keeps every prior month's DM Comments column intact, under its
     original label (e.g. "JUL2026 DM Comments") and appends the new output's 
     own blank comments column (e.g. "AUG2026 DM Comments") after it. Comment
     history accumulates one column per cycle and is not overwritten, so 
     running this each month builds a new column per review cycle.
  2. Highlights in the "Mismatching or Missing SAEs" tab of the merged
     result:
       - YELLOW  = row is new (no matching row existed in the previous output)
       - ORANGE  = row existed before, but a field (including Status)
                   changed since then
     Unhighlighted rows are unchanged from the previous run.

Row identity = Subject Number + Source (EDC/Vendor) + the row's own record
ID (EDC Row # for an EDC-sourced row, Vendor Case # for a vendor-sourced
row).

The "Matching SAEs" tab is carried through from the new output as-is.

Requirements:
    pip install pandas openpyxl
"""

import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# =========================================================================
# CONFIGURATION
# =========================================================================

CONFIG = {
    "previous_output_file": "SAE_Reconciliation_Report_PREVIOUS.xlsx",
    "new_output_file": "SAE_Reconciliation_Report.xlsx",
    "merged_output_file": "SAE_Reconciliation_Report_MERGED.xlsx",

    "matching_sheet": "Matching SAEs",
    "mismatch_sheet": "Mismatching or Missing SAEs",

    "added_fill_color": "FFFF00",    # yellow - new row
    "changed_fill_color": "FFC000",  # orange - existing row, field(s) changed
}


# =========================================================================
# HELPERS
# =========================================================================

def normalize_identity_value(v):
    if pd.isna(v):
        return ""
    s = str(v).strip()
    s = re.sub(r"\.0+$", "", s)  # "104.0" and "104" are congruent
    return s


def row_identity(row, has_source):
    """Anchor identity to the row's own record, not its counterpart's ID
    (which may appear/disappear between runs -- see CONFIG comment)."""
    subject = normalize_identity_value(row.get("Subject Number"))
    if has_source:
        source = normalize_identity_value(row.get("Source"))
        if source == "EDC":
            return ("EDC", subject, normalize_identity_value(row.get("EDC Row #")))
        elif source == "Vendor":
            return ("Vendor", subject, normalize_identity_value(row.get("Vendor Case #")))
    # Fallback if Source column is missing from a file (e.g. an older output).
    # Best effort using whichever ID is present.
    edc_row = normalize_identity_value(row.get("EDC Row #"))
    case_num = normalize_identity_value(row.get("Vendor Case #"))
    if edc_row:
        return ("EDC", subject, edc_row)
    return ("Vendor", subject, case_num)


def find_comments_columns(columns):
    """Comments columns are labeled with the current month, e.g.
    'AUG2026 DM Comments'match on the fixed suffix, in file order.
    A file may have more than one if it's already the product of a
    previous comparison run."""
    return [c for c in columns if str(c).strip().endswith("DM Comments")]


# =========================================================================
# MAIN
# =========================================================================

def wrap_comments_columns(path, sheet_name, column_labels, width=45):
    """Enable wrap-text on the given comment column(s) (by header label) so
    long comments display on multiple lines instead of overflowing, and
    give each a readable width to wrap at."""
    wb = load_workbook(path)
    ws = wb[sheet_name]
    header = [c.value for c in ws[1]]

    for label in column_labels:
        if label not in header:
            continue
        col_idx = header.index(label) + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = width
        for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def main():
    cfg = CONFIG

    prev_df = pd.read_excel(cfg["previous_output_file"], sheet_name=cfg["mismatch_sheet"], dtype=str)
    new_df = pd.read_excel(cfg["new_output_file"], sheet_name=cfg["mismatch_sheet"], dtype=str)

    has_source = "Source" in prev_df.columns and "Source" in new_df.columns
    if not has_source:
        print("WARNING: 'Source' column missing from one of the files -- "
              "falling back to a best-effort identity. Results may be "
              "less precise for rows that gained/lost a counterpart ID "
              "between runs.")

    prev_comment_cols = find_comments_columns(prev_df.columns)
    new_comment_cols = find_comments_columns(new_df.columns)
    if not prev_comment_cols:
        raise ValueError(f"No '... DM Comments' column found in {cfg['previous_output_file']}")
    if not new_comment_cols:
        raise ValueError(f"No '... DM Comments' column found in {cfg['new_output_file']}")

    all_comment_cols = set(prev_comment_cols) | set(new_comment_cols)
    # Columns to check for a changed row. Everything except the
    # run-specific Obs number and any comments column (a comment being 
    # added/edited is not a data discrepancy).
    compare_cols = [c for c in new_df.columns if c not in all_comment_cols and c != "Obs"]

    # Final comment column order: all historical columns first in their
    # original order, then any new-file columns not already present
    # (typically this cycle's blank comment column). If the same label shows
    # up in both files (e.g. reconciliation re-run within the same month
    # before the DM finished the previous review), it's treated as one
    # column. Whatever is in the new file will be maintained, otherwise the
    # old value is carried forward.
    ordered_comment_cols = list(prev_comment_cols)
    for c in new_comment_cols:
        if c not in ordered_comment_cols:
            ordered_comment_cols.append(c)

    prev_by_id = {}
    for _, r in prev_df.iterrows():
        prev_by_id[row_identity(r, has_source)] = r

    row_status = []  # "added" / "changed" / "unchanged", aligned to new_df row order
    carried = {col: [] for col in ordered_comment_cols}

    for _, new_row in new_df.iterrows():
        key = row_identity(new_row, has_source)
        prev_row = prev_by_id.get(key)
        row_status.append("added" if prev_row is None else None)  # status finalized below

        for col in ordered_comment_cols:
            new_val = new_row.get(col) if col in new_df.columns else None
            has_new_val = not (pd.isna(new_val) or str(new_val).strip() == "")
            if has_new_val:
                carried[col].append(new_val)  
            elif prev_row is not None:
                carried[col].append(prev_row.get(col))
            else:
                carried[col].append(None)

        if prev_row is not None:
            changed = any(
                normalize_identity_value(prev_row.get(c)) != normalize_identity_value(new_row.get(c))
                for c in compare_cols
            )
            row_status[-1] = "changed" if changed else "unchanged"

    # Build the merged frame: new_df's content columns, then the comment
    # columns in the order determined above.
    merged_df = new_df.drop(columns=new_comment_cols).copy()
    for col in ordered_comment_cols:
        merged_df[col] = carried[col]

    new_ids = {row_identity(r, has_source) for _, r in new_df.iterrows()}
    resolved_count = sum(1 for k in prev_by_id if k not in new_ids)

    # --- Write merged workbook ---
    with pd.ExcelWriter(cfg["merged_output_file"], engine="openpyxl") as writer:
        try:
            matching_df = pd.read_excel(cfg["new_output_file"], sheet_name=cfg["matching_sheet"], dtype=str)
            matching_df.to_excel(writer, sheet_name=cfg["matching_sheet"], index=False)
        except Exception:
            pass
        merged_df.to_excel(writer, sheet_name=cfg["mismatch_sheet"], index=False)

    # --- Apply highlighting ---
    wb = load_workbook(cfg["merged_output_file"])
    ws = wb[cfg["mismatch_sheet"]]

    added_fill = PatternFill(start_color=cfg["added_fill_color"], end_color=cfg["added_fill_color"], fill_type="solid")
    changed_fill = PatternFill(start_color=cfg["changed_fill_color"], end_color=cfg["changed_fill_color"], fill_type="solid")

    n_cols = merged_df.shape[1]
    for i, status in enumerate(row_status):
        if status == "unchanged":
            continue
        fill = added_fill if status == "added" else changed_fill
        excel_row = i + 2  # +1 for header, +1 for 1-indexing
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=excel_row, column=col_idx).fill = fill

    wb.save(cfg["merged_output_file"])

    wrap_comments_columns(cfg["merged_output_file"], cfg["mismatch_sheet"], ordered_comment_cols)

    print(f"Rows in new output:      {len(merged_df)}")
    print(f"  Added (new):           {row_status.count('added')}  [yellow]")
    print(f"  Changed:               {row_status.count('changed')}  [orange]")
    print(f"  Unchanged:             {row_status.count('unchanged')}")
    print(f"Resolved (in previous, not in new): {resolved_count}")
    print(f"Comment columns carried forward: {prev_comment_cols}")
    print(f"New/current comment column(s): {new_comment_cols}")
    print(f"\nMerged report saved to {cfg['merged_output_file']}")


if __name__ == "__main__":
    main()
